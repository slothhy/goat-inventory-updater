import requests
import os
from openpyxl import load_workbook
from forex_python.converter import CurrencyRates
from datetime import datetime
import json
import time
import collections
import re

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))

sales_db = collections.defaultdict(dict)
slug_db = collections.defaultdict()

def load_config():
    with open(os.path.join(__location__, "config.json")) as json_file:
        data = json.load(json_file)
        return data

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # pylint: disable=no-member
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def main():
    config = load_config()
   
    if config["currency"] == "USD":
        rate = 1 
    else:
        c = CurrencyRates()
        rate = c.get_rate('USD', config["currency"])

    wb = load_workbook(resource_path("./stock_book.xlsx"))
    ws = wb['Sheet1']

    session = requests.Session()
    session.headers.update({
        'content-type': 'application/x-www-form-urlencoded',
        'user-agent': config["user-agent"],
        'accept': '*/*',
        'accept-ending': 'gzip, deflate, br',
        'connection': 'keep-alive'
    })

    row_num = config["start-row"]

    for row in ws.iter_rows(min_row=config["start-row"]): 
        if row[0].value == None:
            break
        print(f'Fetching row {row_num}: {row[0].value}')
        sku = row[0].value
        size = str(row[2].value)
        size = re.sub('[YW]', '', size) #remove Y/W

        if sku in slug_db:
            slug = slug_db[sku]
        else:
            slug = search_product(sku, session)
            if slug is None:
                break
            slug_db[sku] = slug
         
        if sales_db[sku].get(size) is not None:
            print('Skipping, asks exist in database')
            row[8].value = sales_db[sku].get(size)
        else:
            price = get_ask(sku, size, slug, session) / 100
            if price is None:
                break
            row[8].value = price
            sales_db[sku][size] = price

        row_num += 1
        
    wb.save(resource_path("./stock_book_output.xlsx"))

def search_product(sku, session):
    url = "https://2fwotdvm2o-dsn.algolia.net/1/indexes/product_variants_v2_trending_purchase/query?x-algolia-agent=Algolia%20for%20JavaScript%20(3.35.1)%3B%20Browser&x-algolia-application-id=2FWOTDVM2O&x-algolia-api-key=ac96de6fef0e02bb95d433d8d5c7038a"
    payload=f"{{\"params\":\"query=&query={sku}&distinct=true&facetFilters=(product_category%3Ashoes)&page=0&hitsPerPage=1\"}}"

    req = session.post(url, data=payload)
    while req.status_code != 200:
        print(f'Error {req.status_code} at search_product')
        input("Please solve captcha at https://www.goat.com on an incognito window and press enter")
        req = session.post(url, data=payload)
    else:
        data = req.json()
        slug = data["hits"][0]["slug"]
        return slug

def get_ask(sku, size, slug, session):
    url = "https://www.goat.com/web-api/v1/product_variants?productTemplateId=" + slug

    req = session.get(url)
    while req.status_code != 200:
        print(f'Error {req.status_code} at get_sales')
        input("Please solve captcha at https://www.goat.com on an incognito window and press enter")
        req = session.get(url)
    else:
        data = req.json()
        for i in data:
            if str(i["size"]) == size and i["shoeCondition"] == "new_no_defects" and i["boxCondition"] == "good_condition":
                return i["lowestPriceCents"]["amount"]

if __name__ == '__main__':
    main()